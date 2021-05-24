from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl.workbook import Workbook
import time
import openpyxl

alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
northDivision = False

def SetupHeaders(wbActive):
    ### Set up Header for table
    c1 = wbActive.cell(row=1, column=1)
    c1.value = wbActive.title
    c2 = wbActive.cell(row=2, column=1)
    c2.value = "Team:"
    c3 = wbActive['C2']
    c3.value = "GP"
    c4 = wbActive['D2']
    c4.value = "W"
    c5 = wbActive['E2']
    c5.value = "L"
    c6 = wbActive['F2']
    c6.value = "OTL"
    c7 = wbActive['G2']
    c7.value = "Pts"
    c8 = wbActive['H2']
    c8.value = "GF"
    c9 = wbActive['I2']
    c9.value = 'GA'
    c10 = wbActive['J2']
    c10.value = 'Diff'
    c11 = wbActive['K2']
    c11.value = 'L10'
    c12 = wbActive['L2']
    c12.value = 'Strk'


def FillTable(tabNum, northDivision):
    ##set up header
    SetupHeaders(wb.active)
    ### Fill in teams
    count = 2
    for x in range(3, 11):
        cellnum = "A" + str(x)
        TeamCell = wb.active[cellnum]
        if (northDivision and count > 8):
            break
        xpathstr = "/html/body/div[8]/div/div[6]/div/div[2]/span/div/div/div[1]/div/div[2]/div[2]/div/div/div/div[3]/div/div/div/div[2]/div/div/div[2]/div/div[" + str(tabNum) + "]/div/div/table/tbody/tr[" + str(count) + "]/td[3]"
        elemTeam = driver.find_element_by_xpath(xpathstr)
        TeamCell.value = elemTeam.text
        count = count + 1
    ### Team record
    TeamCount = 2
    TeamStatCount = 4
    for x in range(3, 11):
        for y in range(2, 12):
            cellnum = alphabet[y] + str(x)
            TeamStat = wb.active[cellnum]
            if (northDivision and TeamCount > 8):
                break
            xpathstr = "/html/body/div[8]/div/div[6]/div/div[2]/span/div/div/div[1]/div/div[2]/div[2]/div/div/div/div[3]/div/div/div/div[2]/div/div/div[2]/div/div[" + str(tabNum) + "]/div/div/table/tbody/tr[" + str(TeamCount) + "]/td[" + str(TeamStatCount) + "]"
            elemTeamStat = driver.find_element_by_xpath(xpathstr)
            TeamStat.value = elemTeamStat.text

            TeamStatCount = TeamStatCount + 1
        TeamCount = TeamCount + 1
        TeamStatCount = 4

### Webdriver
driver = webdriver.Chrome()
driver.get("https://www.google.com")
assert "Google" in driver.title

### Direct to google search page
que=driver.find_element_by_xpath("//input[@name='q']")
que.send_keys("NHL")
que.send_keys(Keys.RETURN)
time.sleep(1)
elemstandings = driver.find_element_by_class_name("tb_sh")
elemstandings.click()
time.sleep(1)
elemDivision1 = driver.find_element_by_class_name("WUs0Ze")
elemDivision1.find_element_by_xpath("//div[@data-index='0']").click()

### Write to Excell
wb = Workbook()
ws1 = wb.active
##create sheets
ws1.title = "Central Division"
ws2 = wb.create_sheet("Sheet_2")
ws2.title = "East Division"
ws3 = wb.create_sheet("Sheet_3")
ws3.title = "West Division"
ws4 = wb.create_sheet("Sheet_4")
ws4.title = "North Division"

### set active worksheet
wb.active = ws1

FillTable(1, northDivision)

### change active worksheet
wb.active = ws2
### navigate to next tab
elemDivision1 = driver.find_element_by_class_name("WUs0Ze")
elemDivision1.find_element_by_xpath("//div[@data-index='1']").click()

FillTable(2, northDivision)

### change active worksheet
wb.active = ws3
### navigate to next tab
elemDivision1 = driver.find_element_by_class_name("WUs0Ze")
elemDivision1.find_element_by_xpath("//div[@data-index='2']").click()

FillTable(3, northDivision)

### change active worksheet
wb.active = ws4
### navigate to next tab
elemDivision1 = driver.find_element_by_class_name("WUs0Ze")
elemDivision1.find_element_by_xpath("//div[@data-index='3']").click()

northDivision = True
FillTable(4,northDivision)

### save to file
wb.save("D:\\Python\\WebdriverPractice\\openpyxldemo.xlsx")

### Close webdriver
driver.close()
